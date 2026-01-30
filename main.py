import os
import asyncio
import hashlib
import re
from werkzeug.security import generate_password_hash
from array import array
from datetime import datetime, timedelta, date
from functools import wraps
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from pathlib import Path
from src.models.models import db, init_db
from src.models.models import User
from src.api.auth import auth_bp, login_required, roles_required
from src.api.articles import bp as articles_api

#import jwt
#import mysql.connector
import numpy as np
from flask import (Flask, jsonify, make_response, render_template, request,
                   send_from_directory, session, redirect, url_for, render_template, request, redirect, url_for, current_app)
#from flask_sqlalchemy import SQLAlchemy
from numpy import ndarray
from openpyxl import load_workbook

#importacion de algoritmos
from Layout.aco import ejecutar_aco
from Layout.ba import ejecutar_ba
from Layout.da import ejecutar_da
from Layout.daaco import ejecutar_daaco
from Layout.daba import ejecutar_daba
from Layout.dapso import ejecutar_dapso
from Layout.mooraaco import ejecutar_mooraaco
from Layout.mooraba import ejecutar_mooraba
from Layout.moorapso import ejecutar_moorapso
from Layout.moorav import ejecutar_moorav
from Layout.pso import ejecutar_pso
from Layout.topsis import ejecutar_topsis
from Layout.topsisaco import ejecutar_topsisaco
from Layout.topsisba import ejecutar_topsisba
from Layout.topsispso import ejecutar_topsispso

# ----------- BASE DE DATOS: CONFIGURACIÓN ROBUSTA -----------
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_DIR = os.path.join(BASE_DIR, 'db')
DB_PATH = os.path.join(DB_DIR, 'swarm.db')
# print('\n=== DEBUG DE INICIALIZACIÓN DE BASE DE DATOS ===')
# print('Directorio base:', BASE_DIR)
# print('Directorio db:', DB_DIR)
# print('Ruta absoluta de swarm.db:', DB_PATH)

os.makedirs(DB_DIR, exist_ok=True)

# Prueba de permisos de escritura en la carpeta db
try:
    test_path = os.path.join(DB_DIR, 'test.txt')
    with open(test_path, 'w') as f:
        f.write('ok')
    os.remove(test_path)
except Exception as e:
    print('ERROR: No puedo escribir en la carpeta db/:', e)

# ----------- CONFIGURACIÓN DE FLASK Y SQLALCHEMY -----------
app = Flask(__name__, template_folder = 'static/templates')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = '563cebb3aceb49e0a6c79ded5c717235'
user_root = 'Experiments'
today_str = date.today().isoformat()

db.init_app(app)
init_db(app)

app.register_blueprint(auth_bp)
app.register_blueprint(articles_api)

print("DB articles:", id(db))

#-------------------------------------------------------------------------------------------------------------------
@app.route('/acercade')
def acercade():
    return render_template('acercade.html')

@app.route('/casoexperimental')
@roles_required('user','admin', 'superadmin')
def casoexperimental():
    return render_template('casoexperimental.html')
                #Algoritmos PSO
#-------------------------------------------------------------------------------------------------------------------
@app.route('/pso')
@roles_required('user','admin', 'superadmin')
def pso():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))

        return render_template('pso.html', datosPso=datosPso)
    except Exception as e:
        return render_template('pso.html', error_message=str(e))

@app.route('/pso', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_pso():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2, usuario))
        print("Resultados de la ejecución:", datosPso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_pso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/dapso')
@roles_required('user','admin', 'superadmin')
def dapso():
    try:
        # Obtener los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llamar a la función de procesar_datos en pso.py
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))

        return render_template('dapso.html', datosDapso=datosDapso)
    except Exception as e:
        return render_template('dapso.html', error_message=str(e))

@app.route('/dapso', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_dapso():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2, usuario))
        print("Resultados de la ejecución:", datosDapso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDapso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
        print(f'Error en calcular_dapso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------

@app.route('/moorapso', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_moorapso():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input = [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2, usuario))
        print("Resultados de la ejecución:", datosMoorapso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosMoorapso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_moorapso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500


@app.route('/moorapso')
@roles_required('user','admin', 'superadmin')
def moorapso():
    try:
        # Obtén los datos del formulario
        w_input = [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))

        return render_template('moorapso.html', datosMoorapso=datosMoorapso)
    except Exception as e:
        return render_template('moorapso.html', error_message=str(e))
    
    #-------------------------------------------------------------------------------------------------------------------

    #-------------------------------------------------------------------------------------------------------------------

@app.route('/topsispso')
@roles_required('user','admin', 'superadmin')
def topsispso():
     try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('topsispso.html', datosTopsispso=datosTopsispso)
     except Exception as e:
        return render_template('topsispso.html', error_message=str(e))

@app.route('/topsispso', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_topsispso():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
    # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2, usuario))
        print("Resultados de la ejecución:", datosTopsispso)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_topsispso: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------


@app.route('/comparacion')
@roles_required('user','admin', 'superadmin')
def comparacionPura():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('comparacion.html', datosPso=datosPso, datosDapso = datosDapso , datosMoorapso = datosMoorapso, datosTopsispso = datosTopsispso)
    except Exception as e:
        return render_template('comparacion.html', error_message=str(e))


@app.route('/comparacion', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_comparacionPura():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosPso) 
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso, datosDapso, datosMoorapso, datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------


@app.route('/comparacionPso')
@roles_required('user','admin', 'superadmin')
def comparacion():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]
        
        # Llama a la función de procesar_datos en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))

        return render_template('comparacionPso.html', datosPso=datosPso, datosDapso = datosDapso , datosMoorapso = datosMoorapso, datosTopsispso = datosTopsispso)
    except Exception as e:
        return render_template('comparacionPso.html', error_message=str(e))


@app.route('/comparacionPso', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_comparacion():
    try:
        # Obtén los datos del formulario
        w_input = [request.form.get(f'w[{i}]', '') for i in range(5)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        T = int(request.form['T'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Llama a la función de PSO en pso.py
        datosPso = asyncio.run(ejecutar_pso(w, wwi, c1, c2, T, r1, r2))
        datosDapso = asyncio.run(ejecutar_dapso(w, wwi, c1, c2, T, r1, r2))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w, wwi, c1, c2, T, r1, r2))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w, wwi, c1, c2, T, r1, r2))
        print("Resultados de la ejecución:", datosPso) 
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosPso, datosDapso, datosMoorapso, datosTopsispso)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------
            # Algoritmos BA
#-------------------------------------------------------------------------------------------------------------------
@app.route('/ba')
@roles_required('user','admin', 'superadmin')
def ba():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))

        return render_template('ba.html', datosBa=datosBa)
    except Exception as e:
        return render_template('ba.html', error_message=str(e))


@app.route('/ba', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_ba():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max, usuario))
        print("Resultados de la ejecución:", datosBa)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosBa)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_ba: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
            
#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta DA -BA

@app.route('/daba')
@roles_required('user','admin', 'superadmin')
def daba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))

        return render_template('daba.html', datosDaba=datosDaba)
    except Exception as e:
        return render_template('daba.html', error_message=str(e))


@app.route('/daba', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_daba():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max, usuario))
        print("Resultados de la ejecución:", datosDaba)

        # Devuelve los resultados como JSON
        return jsonify(datosDaba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_mooraba: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORA - BA

@app.route('/mooraba')
@roles_required('user','admin', 'superadmin')
def mooraba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))

        return render_template('mooraba.html', datosMooraba=datosMooraba)
    except Exception as e:
        return render_template('mooraba.html', error_message=str(e))


@app.route('/mooraba', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_mooraba():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max, usuario))
        print("Resultados de la ejecución:", datosMooraba)

        # Devuelve los resultados como JSON
        return jsonify(datosMooraba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_mooraba: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/topsisba')
@roles_required('user','admin', 'superadmin')
def topsisba():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))

        return render_template('topsisba.html', datosTopsisba=datosTopsisba)
    except Exception as e:
        return render_template('topsisba.html', error_message=str(e))


@app.route('/topsisba', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_topsisba():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario de la solicitud POST
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max, usuario))
        print("Resultados de la ejecución:", datosTopsisba)

        # Devuelve los resultados como JSON
        return jsonify(datosTopsisba)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_topsisBa: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#Algoritmos ACO

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta DA-ACO

@app.route('/daaco')
@roles_required('user','admin', 'superadmin')
def daaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        
        # Llama a la función de procesar_datos 
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))

        return render_template('daaco.html', datosDaaco=datosDaaco)
    except Exception as e:
        return render_template('daaco.html', error_message=str(e))


@app.route('/daaco', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_daaco():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max, usuario))
        print("Resultados de la ejecución:", datosDaaco)

        # Devuelve los resultados como JSON
        return jsonify(datosDaaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_daaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORA-ACO

@app.route('/mooraaco')
@roles_required('user','admin', 'superadmin')
def mooraaco():
    try:
        ev_input = request.form['ev']  # Obtén el valor
        
        ev_values = ev_input.split(',')
        
        EV = [str(value) for value in ev_values if value.strip() != '']
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos 
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV,w,alphaAco,beta,rho,Q,n_ants,iter_max))

        return render_template('mooraaco.html', datosMooraaco=datosMooraaco)
    except Exception as e:
        return render_template('mooraaco.html', error_message=str(e))


@app.route('/mooraaco', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_mooraaco():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        ev_input = request.form['ev']  # Obtén el valor
        
        ev_values = ev_input.split(',')
        
        EV = [str(value) for value in ev_values if value.strip() != '']
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

        # Llama a la función de PSO en pso.py
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV,w,alphaAco,beta,rho,Q,n_ants,iter_max, usuario))
        print("Resultados de la ejecución:", datosMooraaco)

        # Devuelve los resultados como JSON
        return jsonify(datosMooraaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_Mooraaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
    
    #-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta TOPSIS-ACO

@app.route('/topsisaco')
@roles_required('user','admin', 'superadmin')
def topsisaco():
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # benefit_input = request.form['be']  # Obtén el valor
        
        # benefit_values = benefit_input.split(',')
        
        # benefit_attributes = [int(value) for value in benefit_values if value.strip() != '']
        
        # Llama a la función de procesar_datos 
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w,alphaAco,beta,rho,Q,n_ants,iter_max))

        return render_template('topsisaco.html', datosTopsisaco=datosTopsisaco)
    except Exception as e:
        return render_template('topsisaco.html', error_message=str(e))


@app.route('/topsisaco', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_topsisaco():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # benefit_input = request.form['be']  # Obtén el valor
        
        # benefit_values = benefit_input.split(',')
        
        # benefit_attributes = [int(value) for value in benefit_values if value.strip() != '']

        # Llama a la función de PSO en pso.py
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w,alphaAco,beta,rho,Q,n_ants,iter_max, usuario))
        print("Resultados de la ejecución:", datosTopsisaco)

        # Devuelve los resultados como JSON
        return jsonify(datosTopsisaco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_topsisaco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
@app.route('/comparacionGeneral')
@roles_required('user','admin', 'superadmin')
def comparacionGeneral():
    try:
        # Generales
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos

        #Datos Pso
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Datos Ba
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        

        #Datos Aco
        ev_input = request.form['ev']  # Obtén el valor
        ev_values = ev_input.split(',')
        EV = [str(value) for value in ev_values if value.strip() != '']
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        #Llamar a funciones PSO
        datosDapso = asyncio.run(ejecutar_dapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w,wwi,c1,c2,r1,r2, iter_max))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w,wwi,c1,c2,r1,r2 ,iter_max))
        # Llamar a funciones BA
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
        #Llamar a funciones ACO
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))
        

        return render_template('comparacionGeneral.html', datosDapso=datosDapso, datosMoorapso=datosMoorapso, datosTopsispso=datosTopsispso,
                                datosDaba = datosDaba , datosMooraba = datosMooraba, datosTopsisba = datosTopsisba,
                                datosDaaco=datosDaaco, datosMooraaco=datosMooraaco, datosTopsisaco=datosTopsisaco)
    except Exception as e:
        return render_template('comparacionGeneral.html', error_message=str(e))


@app.route('/comparacionGeneral', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_comparacionGeneral():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
       #Datos Pso
        wwi = float(request.form['wwi'])
        c1 = float(request.form['c1'])
        c2 = float(request.form['c2'])
        # Divide las cadenas de texto en listas
        r1_input = request.form['r1']
        r2_input = request.form['r2']
        r1 = [float(num.strip()) for num in r1_input.split(',')]
        r2 = [float(num.strip()) for num in r2_input.split(',')]

        # Datos Ba
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
    

        #Datos Aco
        ev_input = request.form['ev']  # Obtén el valor
        ev_values = ev_input.split(',')
        EV = [str(value) for value in ev_values if value.strip() != '']
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])

         #Llamar a funciones PSO
        datosDapso = asyncio.run(ejecutar_dapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosMoorapso = asyncio.run(ejecutar_moorapso(w,wwi,c1,c2,r1,r2 ,iter_max))
        datosTopsispso = asyncio.run(ejecutar_topsispso(w,wwi,c1,c2,r1,r2 ,iter_max))
        # Llamar a funciones BA
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
         #Llamar a funciones ACO
        datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
        datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))
        
        
        print("Resultados de la ejecución:", datosDapso) 
        print("Resultados de la ejecución:", datosMoorapso) 
        print("Resultados de la ejecución:", datosTopsispso) 
        print("Resultados de la ejecución:", datosDaba) 
        print("Resultados de la ejecución:", datosMooraba) 
        print("Resultados de la ejecución:", datosTopsisba) 
        print("Resultados de la ejecución:", datosDaaco) 
        print("Resultados de la ejecución:", datosMooraaco) 
        print("Resultados de la ejecución:", datosTopsisaco) 

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDapso,datosMoorapso,datosTopsispso, datosDaba, datosMooraba, datosTopsisba,datosDaaco,datosMooraaco,datosTopsisaco)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------

@app.route('/comparacionBa')
@roles_required('user','admin', 'superadmin')
def comparacionBa():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alpha'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))
        
        
        

        return render_template('comparacionBa.html', datosBa=datosBa, datosDaba = datosDaba , datosMooraba = datosMooraba, datosTopsisba = datosTopsisba)
    except Exception as e:
        return render_template('comparacionBa.html', error_message=str(e))


@app.route('/comparacionBa', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_comparacionBa():
    try:
        # Obtén los datos del formulario
        w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
        alpha = float(request.form['alphaBa'])
        gamma = float(request.form['gamma'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosBa = asyncio.run(ejecutar_ba(w, alpha, gamma, iter_max))
        datosDaba = asyncio.run(ejecutar_daba(w, alpha, gamma, iter_max))
        datosMooraba = asyncio.run(ejecutar_mooraba(w, alpha, gamma, iter_max))
        datosTopsisba = asyncio.run(ejecutar_topsisba(w, alpha, gamma, iter_max))

        # Llama a la función de PSO en pso.py
        
        
        print("Resultados de la ejecución:", datosBa) 
        print("Resultados de la ejecución:", datosDaba) 
        print("Resultados de la ejecución:", datosMooraba) 
        print("Resultados de la ejecución:", datosTopsisba) 

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosBa, datosDaba, datosMooraba, datosTopsisba)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_comparacion: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500

#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
# @app.route('/comparacionAco')
# def comparacionAco():
#     try:
#         # Obtén los datos del formulario
#         w_input =  [float(request.form[f'w{i}']) for i in range(1, 6)]
#         w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
#         alpha = int(request.form['alpha'])
#         beta = int(request.form['beta'])
#         rho = float(request.form['rho'])
#         Q = int(request.form['Q'])
#         n_ants = int(request.form['n_ants'])
#         iter_max = int(request.form['iter_max'])
        
#         # Llama a la función de procesar_datos en pso.py
#         datosAco = asyncio.run(ejecutar_aco(w, alpha, beta, rho, Q, n_ants, iter_max))
#         datosDaaco = asyncio.run(ejecutar_daaco(w, alpha, beta, rho, Q, n_ants, iter_max))
#         datosMooraaco = asyncio.run(ejecutar_mooraaco(w, alpha, beta, iter_max))
#         datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alpha, beta, iter_max))
                

#         return render_template('comparacionAco.html', datosAco=datosAco, datosDaaco = datosDaaco , datosMooraaco = datosMooraaco, datosTopsisaco = datosTopsisaco)
#     except Exception as e:
#         return render_template('comparacionAco.html', error_message=str(e))


@app.route('/comparacionAco', methods=['GET', 'POST'])
@roles_required('user','admin', 'superadmin')
def comparacionAco():
    if request.method == 'POST':
        try:
            ev_input = request.form['ev']  # Obtén el valor
            ev_values = ev_input.split(',')
            EV = [str(value) for value in ev_values if value.strip() != '']

             # Obtén los datos del formulario
            w_input = request.form['w']  # Obtén el valor seleccionado del menú desplegable desde el formulario

            # Divide la cadena en una lista de valores usando la coma como separador
            w_values = w_input.split(',')

            # Convierte cada valor en la lista a un número flotante, filtrando valores vacíos
            w = [float(value) for value in w_values if value.strip() != '']     
            alphaAco = int(request.form['alphaAco'])
            beta = int(request.form['beta'])
            rho = float(request.form['rho'])
            Q = int(request.form['Q'])
            n_ants = int(request.form['n_ants'])
            iter_max = int(request.form['T'])
            
            # Llama a la función de procesar_datos
            datosAco = asyncio.run(ejecutar_aco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosDaaco = asyncio.run(ejecutar_daaco(w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosMooraaco = asyncio.run(ejecutar_mooraaco(EV, w, alphaAco, beta, rho, Q, n_ants, iter_max))
            datosTopsisaco = asyncio.run(ejecutar_topsisaco(w, alphaAco, beta, iter_max))

            # Regresa los resultados como JSON
            return jsonify(datosAco=datosAco, datosDaaco=datosDaaco, datosMooraaco=datosMooraaco, datosTopsisaco=datosTopsisaco)
        except Exception as e:
            # Retorna un error JSON detallado
            return jsonify({'error': str(e)}), 500
    else:
        # Para el método GET, solo renderiza el template
        return render_template('comparacionAco.html')

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmos MCDM - PUROS

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta ACO

@app.route('/aco')
@roles_required('user','admin', 'superadmin')
def aco():
    try:
        # Obtén los datos del formulario
        w = [0.400, 0.200, 0.030, 0.070, 0.300]
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosAco = asyncio.run(ejecutar_aco(w,alphaAco, beta, rho, Q, n_ants, iter_max))

        return render_template('aco.html', datosAco=datosAco)
    except Exception as e:
        return render_template('aco.html', error_message=str(e))


@app.route('/aco', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_aco():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
       # Obtén los datos del formulario
        w = [0.400, 0.200, 0.030, 0.070, 0.300]
        alphaAco = int(request.form['alphaAco'])
        beta = int(request.form['beta'])
        rho = float(request.form['rho'])
        Q = int(request.form['Q'])
        n_ants = int(request.form['n_ants'])
        iter_max = int(request.form['T'])
        
        # Llama a la función de procesar_datos en pso.py
        datosAco = asyncio.run(ejecutar_aco(w, alphaAco, beta, rho, Q, n_ants, iter_max, usuario))
        print("Resultados de la ejecución:", datosAco)

        # Devuelve los resultados como JSON
        return jsonify(datosAco)
    except Exception as e:
        # Manejo de errores
        print(f'Error en calcular_aco: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta TOPSIS

@app.route('/topsis')
@roles_required('user','admin', 'superadmin')
def topsis():
    try:
        # Obtén los datos del formulario
        w_input = float(request.form.get('w', '')) 
        w = [float(value) for value in w_input if value != '']  # Filtra valores vacíos
       
        
        # Llama a la función de procesar_datos en pso.py
        datosTopsis = asyncio.run(ejecutar_topsis(w))

        return render_template('topsis.html', datosTopsis=datosTopsis)
    except Exception as e:
        return render_template('topsis.html', error_message=str(e))


@app.route('/topsis', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_topsis():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante

        # Llama a la función de PSO en pso.py
        datosTopsis = asyncio.run(ejecutar_topsis(w, usuario))
        print("Resultados de la ejecución:", datosTopsis)

        # Devuelve los resultados
        return jsonify(datosTopsis)
    except Exception as e:
        print(f'Error en calcular_topsis: {str(e)}')
        return jsonify({'error': 'Ocurrió un error en el servidor'}), 500


#-------------------------------------------------------------------------------------------------------------------


#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORAV

@app.route('/moorav')
@roles_required('user','admin', 'superadmin')
def moorav():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        
        
        # Llama a la función de procesar_datos en pso.py
        datosMoorav = asyncio.run(ejecutar_moorav(w))

        return render_template('moorav.html', datosMoorav=datosMoorav)
    except Exception as e:
        return render_template('moorav.html', error_message=str(e))


@app.route('/moorav', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_moorav():
    uid = session.get('user_id')  # <-- string key, NO lista
    if uid:
        user = db.session.get(User, uid)   # SQLAlchemy 2.x
        if user:
            usuario = user.username
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        

        # Llama a la función de PSO en pso.py
        datosMoorav = asyncio.run(ejecutar_moorav(w, usuario))
        print("Resultados de la ejecución:", datosMoorav)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosMoorav)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_moorav: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------

#-------------------------------------------------------------------------------------------------------------------
        #Algoritmo_Ruta MOORAV

@app.route('/da')
@roles_required('user','admin', 'superadmin')
def da():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante
        
        
        # Llama a la función de procesar_datos en pso.py
        datosDa = asyncio.run(ejecutar_da(w))

        return render_template('da.html', datosDa=datosDa)
    except Exception as e:
        return render_template('da.html', error_message=str(e))


@app.route('/da', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def calcular_da():
    try:
        # Obtén los datos del formulario
        w_input = request.form.get('w', '')  # Obtiene el valor de 'w' del formulario
        w_values = w_input.split(",")  # Divide la cadena en valores individuales
        w = [float(value.strip()) for value in w_values if value.strip()]  # Convierte cada valor a flotante

        # Llama a la función de PSO en pso.py
        datosDa = asyncio.run(ejecutar_da(w))
        print("Resultados de la ejecución:", datosDa)

        # Obtén los resultados específicos que deseas mostrar
        # dataGBP = resultados['dataGBP']
        # dataGBF = resultados['dataGBF']
        # dataResult = resultados['dataResult']

        # Puedes hacer lo que quieras con los resultados, por ejemplo, pasarlos al template
        return jsonify(datosDa)
    except Exception as e:
        # Manejo de errores, por ejemplo, mostrar un mensaje de error en la interfaz
       print(f'Error en calcular_Da: {str(e)}')
    return jsonify({'error': 'Ocurrió un error en el servidor'}), 500
#-------------------------------------------------------------------------------------------------------------------
@app.route('/index', methods=['POST'])
@roles_required('user','admin', 'superadmin')
def index():
    # Parámetros de control (ingresan)
    w=request.form['w']
    #w=0.3
    # n=int(request.form['n'])  #partículas
    n=9
    c1=request.form['c1']
    #c1=0.50
    c2=request.form['c2']
    #c2=0.50
    e=int(request.form['e'])
    itera=e  #iteraciones

#Función objetivo
# Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
# vector negativo / (vector positivo + vector negarivo)
# R2/(R2+R1)


    # Variables aleatorias, estas van a ser ingresadas(posteriormente)
    #R1 --> es el Vector positivo
    FILE_PATH='r1.xlsx'
    SHEET='Hoja1'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r1=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r1.append(row[0].value)
        #print(r1)

    #R2 --> es el vector negativo
    #SHEET = 'Hoja1'
    FILE_PATH='r2.xlsx'
    workbook=load_workbook(FILE_PATH,read_only=True)
    sheet=workbook[SHEET]
    r2=[]
    for row in sheet.iter_rows():
        #print(row[0].value)
        r2.append(row[0].value)
        #print(r1)

    # ----------
    # Inicialización de la partícula del enjambre
    # Iniciando cálculo de las iteraciones
    print("")
    print("----")
    print("Iteración # 1")
    CP=[]
    for i in range(n):
        r11=float(r1[i])
        CP.append(float(format(10*(r11-0.5), '.4f')))
        #print("CP=",CP)
    #print("   CP=",CP)
    print("   Posición actual:       CP=",CP)

    # Inicialización de la velocidad
    V=[]
    for i in range(n):
        r21=float(r2[i])
        V.append(float(format((r21-0.5), '.4f')))
    #print("   V=",V)
    #print("   Velocidad actual:             V=",V)

    # Óptimo actual CF(1)
    # Ri= (t^-i)/ (t^+i+t^-i) , i=1,...,m
    # vector negativo / (vector positivo + vector negarivo)
    # R2/(R2+R1)
    CF=[]
    for i in range(n):
        #R2/(R1+R2)
        CForm1=float(r1[i])
        CForm2=float(r2[i])
        CForm3=(float((CForm2)/(CForm2+CForm1)))
        CForm=(format(float(CForm3), '.4f'))
        CF.append(float(CForm))
    print("   Óptimo actual:         CF=",CF)
    #print("   CF=",CF)

    # mejor posición actual
    LBP=[]
    for i in range(n):
        # LBP[1]=CP[1]
        LBPP=format(float(CP[i]), '.4f')
        LBP.append(float(LBPP))
    #print("   LBP=",LBP)
    print("   Mejor posición local:  LBP=",LBP)

    # mejor óptimo local
    LBF=[]
    for i in range(n):
        # LBF[1]=CF[1]
        LBFF=format(float(CF[i]), '.4f')
        LBF.append(float(LBFF))
    #print("   LBF=",LBF)
    print("   Mejor óptimo local:    LBF=",LBF)

    # MEJOR ÓPTIMO GLOBAL
    GBF=[]
    GBF.append(max(LBF))
    GBFt=max(LBF)
    #print("   GBF=", GBF)
    print("   Mejor óptimo global:   GBF=", GBF)

    # MEJOR POSICION GLOBAL
    GBP=[]
    GBF_index=LBF.index(GBFt)
    GBPt=float(LBP[GBF_index])
    GBP.append(LBP[GBF_index])
    #print("   GBP=",GBP)
    print("   Mejor posición global: GBP=",GBP)

    ###########################################
    #  Iteráción 2 a la n
    ##########################################
    it=2
    e=e-2

    while (e>=0):
        print("--------")
        print("Iteración #", it)
        long_V1=len(V)-n
        long_LBP1=len(LBP)-n
        long_CP1=len(CP)-n
        long_GBP1=len(GBP)-1
        for i in range(n):
            # V(i+1)= W*V(i) +c1*r1*(LBP(i)-CP(i))+c2*r2*(GBP)i)-CP(i))
            V1=format(float(w)*float(V[long_V1]), '.4f')
            V2=format(float(c1)*float(r1[i]), '.4f')
            V3=format(float(LBP[long_LBP1])-float(CP[long_CP1]), '.4f')
            V4=format(float(c2)*float(r2[i]), '.4f')
            V5=format(float(GBP[long_GBP1]), '.4f')
            V6=format(float(CP[long_CP1]), '.4f')
            V7=float(V5)-float(V6)
            Vx=format((float(V1)+float(V2)*float(V3)+float(V4)*float(V7)), '.4f')
            V.append(Vx)
            long_V1=long_V1+1
            long_LBP1=long_LBP1+1
            long_CP1=long_CP1+1
        long_V2=len(V)-n
        long_CP2=len(CP)-n

        #CP=CP(i)+V(i)
        for i in range(n):
            CPI=(format(float(CP[long_CP2])+float(V[long_V2]), '.4f'))
            CP.append(float(CPI))
            long_V2=long_V2+1
            long_CP2=long_CP2+1

        #óptimo actual CF
        long_CP1=len(CP)-(2*n)
        #print("long_CP2",long_CP2)
        long_CP2=len(CP)-n
        #print("long_CP3",long_CP3)
        for i in range(n):
            #CF2/(CF2+CF1)
            #print("CP",CP)
            CF1=CP[long_CP1]
            #print("CF1",CF1)
            CF2=CP[long_CP2]
            #print("CF2",CF2)
            CF3=float(CF2+CF1)
            CF12=format(float((CF2)/(CF3)),'.4f')
            CF.append(CF12)
            #print("CF12 ",CF12)
            long_CP1=long_CP1+1
            long_CP2=long_CP2+1

        # mejor óptimo local
        long_CF2=len(CF)-n
        long_LBF1=len(LBF)-(n)
        for i in range(n):
            #Max( CF[i],LBF[i-1])
            #print("CF",CF)
            #print("LBF",LBF)

            CFt=float(CF[long_CF2])
            #print("long_CF2",long_CF2)
            #print("CFt",CFt)
            
            LBFt=float(LBF[long_LBF1])
            #print("long_LBF1",long_LBF1)
            #print("LBFt",LBFt)
            if CFt>LBFt:
                LBF.append(CFt)
            else:
                LBF.append(LBFt)
            #print("LBF",LBF)
            long_CF2=long_CF2+1
            long_LBF1=long_LBF1+1

        # mejor posición actual
        long_CP4=len(CP)-n
        for i in range(n):
            #LBP[i]= posición de LBF[i]-CP[i]
            LBPt=(format(float(CP[long_CP4]), '.4f'))
            LBP.append(float(LBPt))
            long_CP4=long_CP4+1

        # MEJOR ÓPTIMO GLOBAL
        long_LBF=len(CP)-n
        temporal=[]
        for i in range(n):
            temporal.append(LBF[long_LBF])
            long_LBF=long_LBF+1
        GBF.append(max(temporal))
        GBFt=max(temporal)
        #print("GBFt=",GBFt)
        #print("MaxGlobal=",GBFt)

        # MEJOR POSICION GLOBAL
        #print("PosiciónGBP posición en CP")
        #print("CP=",CP)
        #long_GBF=len(CP)-n
        long_GBF=len(CP)-1
        #print("LBF",LBF)
        #print("long_CP=",long_GBF)
        GBPt=CP[long_GBF]
        #print("GBPt",GBPt)
        GBP.append(GBPt)


    # impresión de datos por iteración
        V_imp = len(V)-n
        CP_imp = len(CP)-n
        CF_imp = len(CF)-n
        LBF_imp = len(LBF)-n
        LBP_imp = len(LBP)-n
        GBF_imp = len(GBF)-1
        GBP_imp = len(GBP)-1

        V_impT = []
        for i in range(n):
            V_impT.append(float(V[V_imp]))
            V_imp = V_imp+1
        print("   V=",V_impT)

        CP_impT = []
        for i in range(n):
            CP_impT.append(float(CP[CP_imp]))
            CP_imp = CP_imp+1
        print("   CP=",CP_impT)

        CF_impT = []
        for i in range(n):
            CF_impT.append(float(CF[CF_imp]))
            CF_imp = CF_imp+1
        print("   CF=",CF_impT)

        LBF_impT = []
        for i in range(n):
            LBF_impT.append(float(LBF[LBF_imp]))
            LBF_imp = LBF_imp+1
        print("   LBF=",LBF_impT)

        LBP_impT = []
        for i in range(n):
            LBP_impT.append(float(LBP[LBP_imp]))
            LBP_imp = LBP_imp+1
        print("   LBP=",LBP_impT)

        for i in range(n):
            GBF_impT = (float(GBF[GBF_imp]))
        print("   GBF=",GBF_impT)

        for i in range(n):
            GBP_impT = (float(GBP[GBP_imp]))
        print("   GBP=",GBP_impT)

        e = e-1
        it = it+1
        print("")

    print("*******************")
    GBF_SF = len(GBF)-1
    GBP_SF = len(GBP)-1

    GBF_FIN = (float(GBF[GBF_SF]))
    print("   Mejor posición=",GBF_FIN)

    GBP_FIN = (float(GBP[GBP_SF]))
    print("   Mejor óptimo=",GBP_FIN)
    print("*******************")
    print("")

    context = {
        "c1": c1,
        "c2": c2,
        "w": w,
        "n": n,
        "e": e,
        "itera": itera,
        #"var1": var1,
        #"var2": var2,
        #"var3": var3,

        "r1": r1,
        "r2": r2,

        "V_impT": V_impT,
        "CP_impT": CP_impT,
        "CF_impT": CF_impT,
        "LBF_impT": LBF_impT,
        "LBP_impT": LBP_impT,
        "GBF_impT": GBF_impT,
        "GBP_impT": GBP_impT,

        "GBF_FIN": GBF_FIN,
        "GBP_FIN": GBP_FIN,
    }

    return render_template('index.html', **context)

@app.route('/')
@roles_required('user','admin', 'superadmin')
def home():
    uid = session.get('user_id')
    if uid:
        user = db.session.get(User, uid)
        if user:
            usuario = user.username
    return render_template('index.html', usuario=usuario)

def get_username():
    uid = session.get('user_id')
    user = db.session.get(User, uid)
    username = user.username
    
    return username

def obtener_ruta_excel(username, algorithm):
    today_str = datetime.now().strftime("%Y%m%d")
    base_path = "Experiments"
    carpeta = os.path.join(base_path, username, algorithm, today_str)

    if not os.path.isdir(carpeta):
        return None
    
    archivos = os.listdir(carpeta)

    algorithm_files = []

    for archivo in archivos:
        match = re.match(rf"_{algorithm}-(\d+)\.xlsx$", archivo)
        if match:
            timestamp = match.group(1)
            algorithm_files.append((archivo, timestamp))
            
    if not algorithm_files:
        return None
    
    #Se ordena por timestamp
    algorithm_files.sort(key = lambda x: x[1], reverse=True)
    archivo_mas_nuevo = algorithm_files[0][0]
    
    return os.path.join(carpeta, archivo_mas_nuevo)

@app.route('/descargar-pso')
@roles_required('user','admin', 'superadmin')
def descargar_excel_pso():
    username = get_username()
    file = obtener_ruta_excel(username , "PSO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-dapso')
@roles_required('user','admin', 'superadmin')
def descargar_excel_dapso():
    username = get_username()
    file = obtener_ruta_excel(username , "DAPSO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-moorapso')
@roles_required('user','admin', 'superadmin')
def descargar_excel_moorapso():
    username = get_username()
    file = obtener_ruta_excel(username , "MOORAPSO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-topsispso')
@roles_required('user','admin', 'superadmin')
def descargar_excel_topsispso():
    username = get_username()
    file = obtener_ruta_excel(username , "TOPSISPSO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-ba')
@roles_required('user','admin', 'superadmin')
def descargar_excel_ba():
    username = get_username()
    file = obtener_ruta_excel(username , "BA")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-daba')
@roles_required('user','admin', 'superadmin')
def descargar_excel_daba():
    username = get_username()
    file = obtener_ruta_excel(username , "DABA")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-mooraba')
@roles_required('user','admin', 'superadmin')
def descargar_excel_mooraba():
    username = get_username()
    file = obtener_ruta_excel(username , "MOORABA")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-topsisba')
@roles_required('user','admin', 'superadmin')
def descargar_excel_topsisba():
    username = get_username()
    file = obtener_ruta_excel(username , "TOPSISBA")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-aco')
@roles_required('user','admin', 'superadmin')
def descargar_excel_aco():
    username = get_username()
    file = obtener_ruta_excel(username , "ACO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

#Aquí hubo un error
@app.route('/descargar-daaco')
@roles_required('user','admin', 'superadmin')
def descargar_excel_daaco():
    username = get_username()
    file = obtener_ruta_excel(username , "DAACO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-mooraaco')
@roles_required('user','admin', 'superadmin')
def descargar_excel_mooraaco():
    username = get_username()
    file = obtener_ruta_excel(username , "MOORAACO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-topsisaco')
@roles_required('user','admin', 'superadmin')
def descargar_excel_topsisaco():
    username = get_username()
    file = obtener_ruta_excel(username , "TOPSISACO")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-topsis')
@roles_required('user','admin', 'superadmin')
def descargar_excel_topsis():
    username = get_username()
    file = obtener_ruta_excel(username , "TOPSIS")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-moorav')
@roles_required('user','admin', 'superadmin')
def descargar_excel_moorav():
    username = get_username()
    file = obtener_ruta_excel(username , "MOORAV")
    directorio = os.path.dirname(file)
    archivo = os.path.basename(file)
    if not file:
        return "No hay ejecuciones", 404
    
    return send_from_directory(directorio, archivo, as_attachment=True)

@app.route('/descargar-da')
@roles_required('user','admin', 'superadmin')
def descargar_excel_da():
    directorio = 'Experiments/static'  
    filename = 'DA_1.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)

@app.route('/descargar-zip')
@roles_required('user','admin', 'superadmin')
def descargar_zip():
    directorio = 'Experiments/static'  
    filename = 'Compara.zip'
    return send_from_directory(directorio, filename, as_attachment=True)


@app.route('/descargar-parametros')
@roles_required('user','admin', 'superadmin')
def descargar_parametros():
    directorio = 'Experiments/static'  
    filename = 'entradas-Programa.xlsx'
    return send_from_directory(directorio, filename, as_attachment=True)


@app.route('/signup', methods=['GET','POST'])
def signup():
    msg = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()

        # Validaciones mínimas
        if not username:
            msg = 'El nombre de usuario es obligatorio.'
            return render_template('signup.html', msg=msg)
        if len(password) < 4:  # considera subirlo a 8+
            msg = 'La contraseña es demasiado corta.'
            return render_template('signup.html', msg=msg)

        # Usuario único (SQLAlchemy 2.x idiomático)
        stmt = select(User).where(User.username == username)
        existing = db.session.execute(stmt).scalar_one_or_none()
        if existing:
            msg = 'El usuario ya existe.'
            return render_template('signup.html', msg=msg)

        # Crear usuario
        new_user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role='user'
        )

        try:
            db.session.add(new_user)
            db.session.commit()  # ahora new_user.id existe

            # Crear carpeta del usuario: Experiments/<id>-<username-sanitizado>/
            base = Path(current_app.config.get('EXPERIMENTS_ROOT', 'Experiments'))
            dirname = f"{new_user.id}-{secure_filename(new_user.username or 'user')}"
            user_dir = base / dirname
            user_dir.mkdir(parents=True, exist_ok=True)

            # Redirigir a login (ajusta el endpoint si fuera diferente)
            return redirect(url_for('auth.login'))

        except IntegrityError:
            db.session.rollback()
            msg = 'No se pudo crear la cuenta (conflicto de datos).'
        except OSError as e:
            # Si falla la creación de carpeta, puedes optar por seguir o revertir
            msg = f'Cuenta creada, pero no se pudo crear la carpeta del usuario: {e.strerror}'
            # Si prefieres revertir todo:
            # db.session.delete(new_user); db.session.commit()
            # msg = 'Error creando la carpeta del usuario.'

    return render_template('signup.html', msg=msg)

@app.route('/articulos')
@roles_required('admin', 'superadmin')
def articulos():
    directorio = 'Experiments/static'  
    filename = ''
    return render_template('articulos.html')
    
    #return send_from_directory(directorio, filename, as_attachment=True)
@app.route('/publicaciones')
@roles_required('user','admin', 'superadmin')
def publicacion():
    directorio = 'Experiments/static'  
    filename = ''
    return render_template('publicaciones.html')


# @app.route('/descargar-ultimo')
# @roles_required('user','admin', 'superadmin')
# def descargar_ultimo():
#     prefix = request.args.get('prefix')  # opcional: e.g. ?prefix=pso
#     lastf = get_last_file(prefix)
#     if not lastf:
#         return abort(404, description="No hay archivos disponibles aún.")
#     # send_from_directory necesita dir y nombre
#     return send_from_directory(lastf.parent, lastf.name, as_attachment=True)
 
if '__main__' == __name__:
    app.run(port=5000, debug=True)
